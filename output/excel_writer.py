"""
Excel Diary Writer

Per month (written on each run):
  - "Txn - Jan 2026"      — transactions for that month, with Category dropdown
  - "Summary - Jan 2026"  — monthly breakdown, combined totals, category table, charts

Persistent across months:
  - "All-Time Summary"    — rebuilt every run from all Txn sheets in the workbook
  - "_Categories"         — hidden sheet holding the dropdown list (created once)

Append behaviour:
  - If the output file already exists, load it; otherwise create a fresh workbook.
  - If "Txn - MMM YYYY" or "Summary - MMM YYYY" already exist, delete and rewrite.
  - Always delete and rebuild "All-Time Summary" at the end.

Category logic:
  - NEUTRAL_CATEGORIES are excluded from expense totals and net flow.
    They represent money you still own (savings, investments, inter-account moves) — not real consumption.
"""

import os
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import LineChart, Reference

# ── Categories ────────────────────────────────────────────────────────────────

CATEGORIES = [
    "Income",
    "Daily Needs",
    "Lifestyle",
    "Bills & Subscriptions",
    "Shared Costs",
    "Credit Card Payment",
    "Cash Withdrawal",
    "Savings & Investment",
    "Inter-account Transfer",
    "Other",
]

# These categories are excluded from expense totals — money you still own
NEUTRAL_CATEGORIES = {
    "Savings & Investment",
    "Inter-account Transfer",
}

CAT_SHEET = "_Categories"

# ── Palette ───────────────────────────────────────────────────────────────────
C_DARK_BLUE   = "1F4E79"
C_MID_BLUE    = "2E75B6"
C_BCA_ROW     = "DDEEFF"
C_MANDIRI_ROW = "FFF3CD"
C_ALT_ROW     = "F7F7F7"
C_SUBTOTAL_BG = "EBF3FA"
C_WHITE       = "FFFFFF"
C_DEBIT       = "C00000"
C_CREDIT      = "375623"
C_NEUTRAL_CAT = "888888"
C_NEUTRAL_BG  = "EFEFEF"

THIN      = Side(style="thin",   color="CCCCCC")
THICK     = Side(style="medium", color="AAAAAA")
BORDER    = Border(left=THIN,  right=THIN,  top=THIN,  bottom=THIN)
THICK_BOT = Border(left=THIN,  right=THIN,  top=THIN,  bottom=THICK)

IDR_FMT  = '#,##0'
IDR_FMT0 = '#,##0;(#,##0);"-"'
PCT_FMT  = '0.0%'
DATE_FMT = 'DD/MM/YYYY'

MONTH_NAMES = {
    1:"January", 2:"February",  3:"March",    4:"April",
    5:"May",     6:"June",      7:"July",      8:"August",
    9:"September",10:"October",11:"November",12:"December",
}
MONTH_SHORT = {
    "Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,
    "Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12,
}

TXN_HEADERS = [
    "No", "Date", "Bank", "Type", "Category",
    "Description", "Debit (IDR)", "Credit (IDR)", "Balance (IDR)", "Notes"
]
TXN_WIDTHS = {
    "A": 6, "B": 13, "C": 10, "D": 10, "E": 26,
    "F": 50, "G": 18, "H": 18, "I": 18, "J": 30,
}
SUMM_WIDTHS = {
    "A": 28, "B": 10, "C": 18, "D": 18, "E": 18, "F": 16, "G": 22,
    "J": 16, "K": 20, "L": 20, "M": 20,
}


# ── Style helpers ─────────────────────────────────────────────────────────────

def _font(bold=False, color="000000", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def _right():
    return Alignment(horizontal="right", vertical="center")

def _set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def _hdr_cell(ws, row, col, value, bg=C_DARK_BLUE, fg=C_WHITE, size=10, center=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold=True, color=fg, size=size)
    c.fill      = _fill(bg)
    c.border    = BORDER
    c.alignment = _center() if center else _left(wrap=False)
    return c

def _data_cell(ws, row, col, value, bg, fmt=None, bold=False,
               color="000000", align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font   = _font(bold=bold, color=color)
    c.fill   = _fill(bg)
    c.border = BORDER
    if fmt:
        c.number_format = fmt
    if align == "center":
        c.alignment = _center()
    elif align == "right":
        c.alignment = _right()
    else:
        c.alignment = _left()
    return c

def _section_header(ws, row, value, cols=7):
    _hdr_cell(ws, row, 1, value, bg=C_DARK_BLUE, size=11)
    ws.merge_cells(f"A{row}:{get_column_letter(cols)}{row}")


# ── Transactions sheet ────────────────────────────────────────────────────────

def _write_txn_sheet(ws, transactions):
    ws.freeze_panes = "A2"
    cat_range = f"'{CAT_SHEET}'!$A$1:$A${len(CATEGORIES)}"

    for ci, hdr in enumerate(TXN_HEADERS, start=1):
        _hdr_cell(ws, 1, ci, hdr)
    ws.row_dimensions[1].height = 18

    for ri, txn in enumerate(transactions, start=2):
        bank     = txn.get("bank", "")
        bg       = C_BCA_ROW if bank == "BCA" else (C_MANDIRI_ROW if bank == "Mandiri" else C_ALT_ROW)
        txn_type = txn.get("type", "")
        _data_cell(ws, ri, 1, ri - 1,                    bg, align="center")
        _data_cell(ws, ri, 2, txn.get("date"),           bg, fmt=DATE_FMT, align="center")
        _data_cell(ws, ri, 3, bank,                      bg, bold=True, align="center")
        _data_cell(ws, ri, 4, txn_type,                  bg, align="center",
                   color=C_DEBIT if txn_type == "Debit" else (C_CREDIT if txn_type == "Credit" else "000000"))
        _data_cell(ws, ri, 5, txn.get("category", ""),  bg, align="left")
        c = _data_cell(ws, ri, 6, txn.get("description", ""), bg, align="left")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        _data_cell(ws, ri, 7, txn.get("debit"),          bg, fmt=IDR_FMT0, align="right",
                   color=C_DEBIT if txn.get("debit") else "000000")
        _data_cell(ws, ri, 8, txn.get("credit"),         bg, fmt=IDR_FMT0, align="right",
                   color=C_CREDIT if txn.get("credit") else "000000")
        _data_cell(ws, ri, 9, txn.get("balance"),        bg, fmt=IDR_FMT0, align="right")
        _data_cell(ws, ri, 10, txn.get("notes", ""),     bg, align="left")

    last  = len(transactions) + 1
    total = last + 1
    for ci in range(1, 11):
        c        = ws.cell(row=total, column=ci)
        c.fill   = _fill(C_DARK_BLUE)
        c.border = THICK_BOT
        c.font   = _font(bold=True, color=C_WHITE)
    ws.cell(row=total, column=6, value="TOTAL").alignment = _right()
    for ci, col in [(7, "G"), (8, "H")]:
        c               = ws.cell(row=total, column=ci, value=f"=SUM({col}2:{col}{last})")
        c.number_format = IDR_FMT
        c.alignment     = _right()

    dv = DataValidation(type="list", formula1=cat_range,
                        allow_blank=True, showErrorMessage=False)
    dv.sqref = f"E2:E{last}"
    ws.add_data_validation(dv)
    _set_widths(ws, TXN_WIDTHS)


# ── Summary sheet (per month) ─────────────────────────────────────────────────

def _write_summary_sheet(ws, transactions, txn_sheet_name):
    agg = defaultdict(lambda: {"count": 0, "debit": 0.0, "credit": 0.0,
                                "last_balance": None, "last_date": None})
    for txn in transactions:
        d = txn.get("date")
        if not d:
            continue
        key = (d.year, d.month, txn.get("bank", "Unknown"))
        g   = agg[key]
        g["count"]  += 1
        g["debit"]  += txn.get("debit")  or 0.0
        g["credit"] += txn.get("credit") or 0.0
        if txn.get("balance") is not None:
            if g["last_date"] is None or d >= g["last_date"]:
                g["last_date"]    = d
                g["last_balance"] = txn["balance"]

    sorted_keys = sorted(agg.keys())
    periods     = sorted(set((y, m) for y, m, _ in sorted_keys))
    banks       = sorted(set(b for _, _, b in sorted_keys))
    row = 1

    # Section 1: Per-bank breakdown
    _section_header(ws, row, "MONTHLY BREAKDOWN BY BANK")
    row += 1
    for ci, h in enumerate(["Period", "Bank", "# Txn", "Total Debit (IDR)",
                             "Total Credit (IDR)", "Net Flow (IDR)", "Closing Balance (IDR)"], 1):
        _hdr_cell(ws, row, ci, h, bg=C_MID_BLUE)
    row += 1
    for (year, month, bank) in sorted_keys:
        g   = agg[(year, month, bank)]
        bg  = C_BCA_ROW if bank == "BCA" else (C_MANDIRI_ROW if bank == "Mandiri" else C_ALT_ROW)
        net = g["credit"] - g["debit"]
        _data_cell(ws, row, 1, f"{MONTH_NAMES[month]} {year}", bg, align="center")
        _data_cell(ws, row, 2, bank,              bg, bold=True, align="center")
        _data_cell(ws, row, 3, g["count"],        bg, align="center")
        _data_cell(ws, row, 4, g["debit"] or None,   bg, fmt=IDR_FMT0, align="right", color=C_DEBIT)
        _data_cell(ws, row, 5, g["credit"] or None,  bg, fmt=IDR_FMT0, align="right", color=C_CREDIT)
        _data_cell(ws, row, 6, net,               bg, fmt=IDR_FMT0, align="right",
                   color=C_CREDIT if net >= 0 else C_DEBIT, bold=True)
        _data_cell(ws, row, 7, g["last_balance"], bg, fmt=IDR_FMT0, align="right")
        row += 1

    # Section 2: Combined totals (excluding neutral categories)
    row += 1
    _section_header(ws, row, "COMBINED TOTALS PER PERIOD  (excludes Savings, Transfers & CC Payments)")
    row += 1
    for ci, h in enumerate(["Period", "Total Txn", "Real Spending (IDR)", "Total Credit (IDR)",
                             "Net Flow (IDR)", "Income/Expense Ratio"], 1):
        _hdr_cell(ws, row, ci, h, bg=C_MID_BLUE)
    ws.merge_cells(f"F{row}:G{row}")
    row += 1

    # Build neutral category exclusion formula fragments
    neutral_debit_excl  = "".join(
        f"-SUMIFS('{txn_sheet_name}'!G:G,'{txn_sheet_name}'!E:E,\"{cat}\")"
        for cat in NEUTRAL_CATEGORIES
    )
    neutral_credit_excl = "".join(
        f"-SUMIFS('{txn_sheet_name}'!H:H,'{txn_sheet_name}'!E:E,\"{cat}\")"
        for cat in NEUTRAL_CATEGORIES
    )

    for (year, month) in periods:
        td    = sum(agg[(year, month, b)]["debit"]  for b in banks if (year, month, b) in agg)
        tc    = sum(agg[(year, month, b)]["credit"] for b in banks if (year, month, b) in agg)
        tn    = sum(agg[(year, month, b)]["count"]  for b in banks if (year, month, b) in agg)
        net   = tc - td
        ratio = (tc / td) if td else None

        # Real spending uses formulas so they recalc when categories are filled in
        real_spend_formula  = f"=SUMIF('{txn_sheet_name}'!G:G,\">0\")"+neutral_debit_excl
        real_credit_formula = f"=SUMIF('{txn_sheet_name}'!H:H,\">0\")"+neutral_credit_excl

        _data_cell(ws, row, 1, f"{MONTH_NAMES[month]} {year}", C_SUBTOTAL_BG, align="center", bold=True)
        _data_cell(ws, row, 2, tn, C_SUBTOTAL_BG, align="center")

        c = _data_cell(ws, row, 3, real_spend_formula,  C_SUBTOTAL_BG, fmt=IDR_FMT0, align="right", color=C_DEBIT, bold=True)
        c = _data_cell(ws, row, 4, real_credit_formula, C_SUBTOTAL_BG, fmt=IDR_FMT0, align="right", color=C_CREDIT, bold=True)

        _data_cell(ws, row, 5, net, C_SUBTOTAL_BG, fmt=IDR_FMT0, align="right",
                   color=C_CREDIT if net >= 0 else C_DEBIT, bold=True)
        _data_cell(ws, row, 6, ratio, C_SUBTOTAL_BG, fmt=PCT_FMT, align="center",
                   color=C_CREDIT if (ratio and ratio >= 1) else C_DEBIT)
        ws.merge_cells(f"F{row}:G{row}")
        row += 1

    # Section 3: Category breakdown with budget
    row += 2
    _section_header(ws, row, "SPENDING BY CATEGORY", cols=5)
    row += 1
    note = ws.cell(row=row, column=1,
                   value="⚑  Fill the Category column in the Transactions sheet — this table updates automatically.  "
                         "Neutral categories (grey) are excluded from real spending totals.")
    note.font      = _font(italic=True, color="595959", size=9)
    note.alignment = _left(wrap=False)
    ws.merge_cells(f"A{row}:E{row}")
    row += 1

    cat_hdr_row = row
    for ci, h in enumerate(["Category", "# Txn", "Debit (IDR)", "Credit (IDR)", "Net (IDR)"], 1):
        _hdr_cell(ws, row, ci, h, bg=C_MID_BLUE)
    row += 1

    cat_start = row
    for cat in CATEGORIES:
        is_neutral = cat in NEUTRAL_CATEGORIES
        bg    = C_NEUTRAL_BG if is_neutral else (C_ALT_ROW if (row % 2 == 0) else C_WHITE)
        color = C_NEUTRAL_CAT if is_neutral else "000000"

        _data_cell(ws, row, 1, f"{cat}  ·  neutral" if is_neutral else cat,
                   bg, align="left", color=color)
        _data_cell(ws, row, 2,
                   f"=COUNTIFS('{txn_sheet_name}'!E:E,\"{cat}\")",
                   bg, align="center", color=color)
        _data_cell(ws, row, 3,
                   f"=SUMIFS('{txn_sheet_name}'!G:G,'{txn_sheet_name}'!E:E,\"{cat}\")",
                   bg, fmt=IDR_FMT0, align="right",
                   color=C_NEUTRAL_CAT if is_neutral else C_DEBIT)
        _data_cell(ws, row, 4,
                   f"=SUMIFS('{txn_sheet_name}'!H:H,'{txn_sheet_name}'!E:E,\"{cat}\")",
                   bg, fmt=IDR_FMT0, align="right",
                   color=C_NEUTRAL_CAT if is_neutral else C_CREDIT)
        _data_cell(ws, row, 5, f"=D{row}-C{row}",
                   bg, fmt=IDR_FMT0, align="right", color=color)
        row += 1

    cat_end = row - 1

    # Totals row (real spending only — exclude neutral)
    _data_cell(ws, row, 1, "REAL SPENDING TOTAL", C_DARK_BLUE, bold=True, color=C_WHITE, align="right")

    non_neutral_rows = [
        str(r) for r, cat in zip(range(cat_start, cat_end + 1), CATEGORIES)
        if cat not in NEUTRAL_CATEGORIES
    ]
    for ci, col in [(2, "B"), (3, "C"), (4, "D"), (5, "E")]:
        formula = f"=SUM({','.join(f'{col}{r}' for r in non_neutral_rows)})" if non_neutral_rows else 0
        c               = ws.cell(row=row, column=ci, value=formula)
        c.fill          = _fill(C_DARK_BLUE)
        c.font          = _font(bold=True, color=C_WHITE)
        c.border        = THICK_BOT
        c.number_format = IDR_FMT
        c.alignment     = _right() if ci > 1 else _center()

    row += 2

    _set_widths(ws, SUMM_WIDTHS)
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 24
    _write_charts(ws, transactions, row)


# ── Charts ────────────────────────────────────────────────────────────────────

def _write_charts(ws, transactions, anchor_row, col_start=10):
    balance_by_date_bank = {}
    for txn in sorted(transactions, key=lambda t: (t.get("date") or "", t.get("bank", ""))):
        if txn.get("date") and txn.get("balance") is not None:
            balance_by_date_bank[(txn["date"], txn["bank"])] = txn["balance"]

    _hdr_cell(ws, 1, col_start, "CHART DATA", bg=C_DARK_BLUE, size=11)
    ws.merge_cells(start_row=1, start_column=col_start,
                   end_row=1,   end_column=col_start + 3)

    hdr_row  = 2
    data_row = 3
    for ci, label in enumerate(["Date", "BCA Balance", "Mandiri Balance", "Combined Balance"]):
        _hdr_cell(ws, hdr_row, col_start + ci, label, bg=C_MID_BLUE)

    chart_dates = sorted(set(d for d, _ in balance_by_date_bank))
    last_bal = {"BCA": None, "Mandiri": None}
    for d in chart_dates:
        for bank in ["BCA", "Mandiri"]:
            if (d, bank) in balance_by_date_bank:
                last_bal[bank] = balance_by_date_bank[(d, bank)]
        bca_b  = last_bal["BCA"]
        man_b  = last_bal["Mandiri"]
        comb_b = (bca_b or 0) + (man_b or 0) if (bca_b is not None or man_b is not None) else None
        ws.cell(row=data_row, column=col_start,   value=d).number_format = DATE_FMT
        ws.cell(row=data_row, column=col_start+1, value=bca_b)
        ws.cell(row=data_row, column=col_start+2, value=man_b)
        ws.cell(row=data_row, column=col_start+3, value=comb_b)
        data_row += 1

    max_data_row = data_row - 1

    def _chart(title, col_offsets, anchor):
        chart               = LineChart()
        chart.title         = title
        chart.style         = 10
        chart.height        = 12
        chart.width         = 22
        chart.y_axis.title  = "Balance (IDR)"
        chart.x_axis.title  = "Date"
        chart.y_axis.numFmt = "#,##0"
        dates = Reference(ws, min_col=col_start, min_row=3, max_row=max_data_row)
        for offset in col_offsets:
            chart.add_data(Reference(ws, min_col=col_start + offset,
                                     min_row=hdr_row, max_row=max_data_row),
                           titles_from_data=True)
        chart.set_categories(dates)
        ws.add_chart(chart, anchor)

    _chart("Combined Balance Over Time", [3],    f"A{anchor_row}")
    _chart("Balance by Bank Over Time",  [1, 2], f"A{anchor_row + 25}")


# ── All-Time Category Section ─────────────────────────────────────────────────

def _write_alltime_category_section(ws, wb, row):
    txn_sheets = sorted(
        [(key, name) for name in wb.sheetnames if (key := _sheet_month_key(name))],
        key=lambda x: x[0]
    )
    if not txn_sheets:
        return

    # Show all categories except Inter-account Transfer — Savings & Investment
    # is neutral for spending totals but shown here for visibility
    HIDDEN_FROM_TABLE = NEUTRAL_CATEGORIES - {"Savings & Investment"}
    display_cats = [c for c in CATEGORIES if c not in HIDDEN_FROM_TABLE]
    n_cols = 1 + len(display_cats)

    _section_header(ws, row, "ALL-TIME BY CATEGORY  (Inter-account Transfers excluded)", cols=n_cols)
    row += 1

    _hdr_cell(ws, row, 1, "Period", bg=C_MID_BLUE)
    for ci, cat in enumerate(display_cats):
        _hdr_cell(ws, row, 2 + ci, cat, bg=C_MID_BLUE)
    row += 1

    data_start = row
    for (year, month), sheet_name in txn_sheets:
        bg = C_ALT_ROW if (row % 2 == 0) else C_WHITE
        _data_cell(ws, row, 1, f"{MONTH_NAMES[month]} {year}", bg, align="center", bold=True)
        for ci, cat in enumerate(display_cats):
            col    = 2 + ci
            credit = f"SUMIFS('{sheet_name}'!H:H,'{sheet_name}'!E:E,\"{cat}\")"
            debit  = f"SUMIFS('{sheet_name}'!G:G,'{sheet_name}'!E:E,\"{cat}\")"
            c = _data_cell(ws, row, col, f"={credit}-{debit}", bg, fmt=IDR_FMT0, align="right")
            c.number_format = '#,##0;[Red]-#,##0;"-"'
        row += 1

    data_end = row - 1

    _data_cell(ws, row, 1, "TOTAL", C_DARK_BLUE, bold=True, color=C_WHITE, align="right")
    for ci in range(len(display_cats)):
        col        = 2 + ci
        col_letter = get_column_letter(col)
        c               = ws.cell(row=row, column=col,
                            value=f"=SUM({col_letter}{data_start}:{col_letter}{data_end})")
        c.fill          = _fill(C_DARK_BLUE)
        c.font          = _font(bold=True, color=C_WHITE)
        c.border        = THICK_BOT
        c.number_format = IDR_FMT
        c.alignment     = _right()

    ws.column_dimensions["A"].width = 22
    for ci in range(len(display_cats)):
        ws.column_dimensions[get_column_letter(2 + ci)].width = 22


# ── All-Time Summary ──────────────────────────────────────────────────────────

def _write_alltime_sheet(ws, wb, all_transactions):
    agg = defaultdict(lambda: {"count": 0, "debit": 0.0, "credit": 0.0,
                                "last_balance": None, "last_date": None})
    for txn in all_transactions:
        d = txn.get("date")
        if not d:
            continue
        key = (d.year, d.month, txn.get("bank", "Unknown"))
        g   = agg[key]
        g["count"]  += 1
        g["debit"]  += txn.get("debit")  or 0.0
        g["credit"] += txn.get("credit") or 0.0
        if txn.get("balance") is not None:
            if g["last_date"] is None or d >= g["last_date"]:
                g["last_date"]    = d
                g["last_balance"] = txn["balance"]

    sorted_keys = sorted(agg.keys())
    periods     = sorted(set((y, m) for y, m, _ in sorted_keys))
    banks       = sorted(set(b for _, _, b in sorted_keys))
    row = 1

    # Section 1: Per-bank per-month
    _section_header(ws, row, "ALL MONTHS — BREAKDOWN BY BANK")
    row += 1
    for ci, h in enumerate(["Period", "Bank", "# Txn", "Total Debit (IDR)",
                             "Total Credit (IDR)", "Net Flow (IDR)", "Closing Balance (IDR)"], 1):
        _hdr_cell(ws, row, ci, h, bg=C_MID_BLUE)
    row += 1
    for (year, month, bank) in sorted_keys:
        g   = agg[(year, month, bank)]
        bg  = C_BCA_ROW if bank == "BCA" else (C_MANDIRI_ROW if bank == "Mandiri" else C_ALT_ROW)
        net = g["credit"] - g["debit"]
        _data_cell(ws, row, 1, f"{MONTH_NAMES[month]} {year}", bg, align="center")
        _data_cell(ws, row, 2, bank,              bg, bold=True, align="center")
        _data_cell(ws, row, 3, g["count"],        bg, align="center")
        _data_cell(ws, row, 4, g["debit"] or None,   bg, fmt=IDR_FMT0, align="right", color=C_DEBIT)
        _data_cell(ws, row, 5, g["credit"] or None,  bg, fmt=IDR_FMT0, align="right", color=C_CREDIT)
        _data_cell(ws, row, 6, net,               bg, fmt=IDR_FMT0, align="right",
                   color=C_CREDIT if net >= 0 else C_DEBIT, bold=True)
        _data_cell(ws, row, 7, g["last_balance"], bg, fmt=IDR_FMT0, align="right")
        row += 1

    # Section 2: Combined totals per period
    row += 1
    _section_header(ws, row, "COMBINED TOTALS PER PERIOD")
    row += 1
    for ci, h in enumerate(["Period", "Total Txn", "Total Debit (IDR)", "Total Credit (IDR)",
                             "Net Flow (IDR)", "Income/Expense Ratio"], 1):
        _hdr_cell(ws, row, ci, h, bg=C_MID_BLUE)
    ws.merge_cells(f"F{row}:G{row}")
    row += 1
    for (year, month) in periods:
        td    = sum(agg[(year, month, b)]["debit"]  for b in banks if (year, month, b) in agg)
        tc    = sum(agg[(year, month, b)]["credit"] for b in banks if (year, month, b) in agg)
        tn    = sum(agg[(year, month, b)]["count"]  for b in banks if (year, month, b) in agg)
        net   = tc - td
        ratio = (tc / td) if td else None
        _data_cell(ws, row, 1, f"{MONTH_NAMES[month]} {year}", C_SUBTOTAL_BG, align="center", bold=True)
        _data_cell(ws, row, 2, tn,         C_SUBTOTAL_BG, align="center")
        _data_cell(ws, row, 3, td or None, C_SUBTOTAL_BG, fmt=IDR_FMT0, align="right", color=C_DEBIT,  bold=True)
        _data_cell(ws, row, 4, tc or None, C_SUBTOTAL_BG, fmt=IDR_FMT0, align="right", color=C_CREDIT, bold=True)
        _data_cell(ws, row, 5, net,        C_SUBTOTAL_BG, fmt=IDR_FMT0, align="right",
                   color=C_CREDIT if net >= 0 else C_DEBIT, bold=True)
        _data_cell(ws, row, 6, ratio,      C_SUBTOTAL_BG, fmt=PCT_FMT, align="center",
                   color=C_CREDIT if (ratio and ratio >= 1) else C_DEBIT)
        ws.merge_cells(f"F{row}:G{row}")
        row += 1

    # Section 3: Grand total
    row += 1
    _section_header(ws, row, "GRAND TOTAL — ALL TIME")
    row += 1
    for ci, h in enumerate(["", "Total Txn", "Total Debit (IDR)", "Total Credit (IDR)",
                             "Net Flow (IDR)", "Income/Expense Ratio"], 1):
        _hdr_cell(ws, row, ci, h, bg=C_MID_BLUE)
    ws.merge_cells(f"F{row}:G{row}")
    row += 1
    gd     = sum(g["debit"]  for g in agg.values())
    gc     = sum(g["credit"] for g in agg.values())
    gn     = sum(g["count"]  for g in agg.values())
    gnet   = gc - gd
    gratio = (gc / gd) if gd else None
    _data_cell(ws, row, 1, "All Months", C_DARK_BLUE, bold=True, color=C_WHITE, align="center")
    _data_cell(ws, row, 2, gn,          C_DARK_BLUE, bold=True, color=C_WHITE,  align="center")
    _data_cell(ws, row, 3, gd or None,  C_DARK_BLUE, bold=True, color="FFCCCC", fmt=IDR_FMT0, align="right")
    _data_cell(ws, row, 4, gc or None,  C_DARK_BLUE, bold=True, color="CCFFCC", fmt=IDR_FMT0, align="right")
    _data_cell(ws, row, 5, gnet,        C_DARK_BLUE, bold=True,
               color="CCFFCC" if gnet >= 0 else "FFCCCC", fmt=IDR_FMT0, align="right")
    _data_cell(ws, row, 6, gratio,      C_DARK_BLUE, bold=True,
               color="CCFFCC" if (gratio and gratio >= 1) else "FFCCCC",
               fmt=PCT_FMT, align="center")
    ws.merge_cells(f"F{row}:G{row}")
    row += 2

    # Section 4: Category breakdown across all months (neutral excluded)
    _write_alltime_category_section(ws, wb, row)

    ws.column_dimensions["A"].width = 28
    for ci in range(1, 8):
        ws.column_dimensions[get_column_letter(1 + ci)].width = 22

    display_cats    = [c for c in CATEGORIES if c not in NEUTRAL_CATEGORIES]
    chart_col_start = 2 + len(display_cats) + 2
    _write_charts(ws, all_transactions, row + len(display_cats) + 5, col_start=chart_col_start)


# ── Read existing Txn sheet ───────────────────────────────────────────────────

def _read_txn_sheet(ws):
    transactions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 9:
            continue
        _, dt, bank, txn_type, category, desc, debit, credit, balance = row[:9]
        notes = row[9] if len(row) > 9 else ""
        if bank not in ("BCA", "Mandiri"):
            continue
        if dt is None:
            continue
        d = dt.date() if hasattr(dt, "date") else dt
        transactions.append({
            "date":        d,
            "bank":        bank,
            "type":        txn_type,
            "category":    category or "",
            "description": desc or "",
            "debit":       float(debit)   if debit   is not None else None,
            "credit":      float(credit)  if credit  is not None else None,
            "balance":     float(balance) if balance is not None else None,
            "notes":       notes or "",
        })
    return transactions


def _sheet_month_key(title):
    """'Txn - Jan 2026' → (2026, 1) or None."""
    if not title.startswith("Txn - "):
        return None
    parts = title[len("Txn - "):].split()
    if len(parts) != 2:
        return None
    month = MONTH_SHORT.get(parts[0])
    try:
        year = int(parts[1])
    except ValueError:
        return None
    return (year, month) if month else None


# ── Entry point ───────────────────────────────────────────────────────────────

def write_diary(transactions, output_path):
    dates = [t["date"] for t in transactions if t.get("date")]
    if not dates:
        raise ValueError("No dated transactions provided.")

    first_date      = min(dates)
    month_abbr      = list(MONTH_SHORT.keys())[first_date.month - 1]
    month_label     = f"{month_abbr} {first_date.year}"
    txn_sheet_name  = f"Txn - {month_label}"
    summ_sheet_name = f"Summary - {month_label}"

    if os.path.exists(output_path):
        wb = load_workbook(output_path)
        for name in wb.sheetnames:
            if not name.startswith("Txn - "):
                continue
            ws = wb[name]
            for row in ws.iter_rows(min_row=2, values_only=False):
                ws.row_dimensions[row[0].row].height = None
            _set_widths(ws, TXN_WIDTHS)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Ensure _Categories hidden sheet exists (recreate if categories changed)
    if CAT_SHEET in wb.sheetnames:
        del wb[CAT_SHEET]
    cat_ws = wb.create_sheet(CAT_SHEET)
    cat_ws.sheet_state = "hidden"
    for i, cat in enumerate(CATEGORIES, start=1):
        cat_ws.cell(row=i, column=1, value=cat)

    for name in [txn_sheet_name, summ_sheet_name, "All-Time Summary"]:
        if name in wb.sheetnames:
            del wb[name]

    existing = sorted(
        [(key, name) for name in wb.sheetnames if (key := _sheet_month_key(name))],
        key=lambda x: x[0]
    )
    new_key    = (first_date.year, first_date.month)
    insert_pos = 0
    for key, name in existing:
        if key < new_key:
            idx = wb.sheetnames.index(name)
            summ_name = name.replace("Txn - ", "Summary - ")
            if summ_name in wb.sheetnames:
                insert_pos = wb.sheetnames.index(summ_name) + 1
            else:
                insert_pos = idx + 1

    ws_txn = wb.create_sheet(txn_sheet_name, insert_pos)
    _write_txn_sheet(ws_txn, transactions)

    ws_summ = wb.create_sheet(summ_sheet_name, insert_pos + 1)
    _write_summary_sheet(ws_summ, transactions, txn_sheet_name)

    all_transactions = list(transactions)
    for name in wb.sheetnames:
        if name.startswith("Txn - ") and name != txn_sheet_name:
            all_transactions.extend(_read_txn_sheet(wb[name]))
    all_transactions.sort(key=lambda t: t["date"])

    cat_pos = wb.sheetnames.index(CAT_SHEET)
    ws_all  = wb.create_sheet("All-Time Summary", cat_pos)
    _write_alltime_sheet(ws_all, wb, all_transactions)

    wb.save(output_path)